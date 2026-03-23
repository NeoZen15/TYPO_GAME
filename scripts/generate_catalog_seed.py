"""
generate_catalog_seed.py

Generate canonical catalog seed JSON files inside `content/catalog/` from the
current source Excel and font manifest.

Usage:
    .venv/bin/python scripts/generate_catalog_seed.py
    .venv/bin/python scripts/generate_catalog_seed.py \
      --excel /path/to/catalogue.xlsx \
      --manifest content/typefaces/font-manifest-v4.json \
      --output-dir content/catalog
"""

from __future__ import annotations

import argparse
import hashlib
import json
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_EXCEL = Path(
    "/Users/launaymarion/Documents/JEUX_DE_TYPO/02_TYPO_ASSETS/06_catalogue/JEUX_DE_TYPO_catalogue_v4__VALIDE__20260219_1903.xlsx"
)
DEFAULT_MANIFEST = Path("content/typefaces/font-manifest-v4.json")
DEFAULT_OUTPUT_DIR = Path("content/catalog")
DEFAULT_SHEET = "typefaces"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate canonical catalog seed JSON files.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL))
    parser.add_argument("--manifest", default=str(DEFAULT_MANIFEST))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    return parser.parse_args()


def normalize_ascii(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    normalized = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return "".join(ch for ch in normalized.lower() if ch.isalnum())


def to_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_excel_rows(excel_path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook[sheet_name]
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell) for cell in header_cells]

    rows: list[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def load_manifest(manifest_path: Path) -> dict:
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def build_typeface_records(excel_rows: list[dict]) -> list[dict]:
    records: list[dict] = []

    for row in excel_rows:
        structural_signature = row.get("structural_signature")
        if isinstance(structural_signature, str):
            structural_signature = json.loads(structural_signature)

        notes = [
            "seeded_from_original_excel_v4",
            "license/designer/foundry metadata still requires enrichment outside the original v4 source",
        ]
        if str(row.get("font_source", "")).strip().lower() == "local":
            notes.append("local/system font requires OFL-compatible replacement before production launch")

        records.append(
            {
                "typeface_slug": row["typeface_slug"],
                "display_name": row["display_name"],
                "display_name_ascii": normalize_ascii(str(row["display_name"])),
                "primary_category": row["primary_category"],
                "sub_category": row["sub_category"],
                "visual_cluster_id": row["visual_cluster_id"],
                "dreyfus_tier": row["dreyfus_tier"],
                "difficulty_base": row["difficulty_base"],
                "rarity_tag": row["rarity_tag"],
                "activation_status": to_bool(row.get("activation_status")),
                "font_source": row["font_source"],
                "is_variable_font": to_bool(row.get("is_variable_font")),
                "year_tag": row["year_tag"],
                "weight_structure": row["weight_structure"],
                "contrast_profile": row["contrast_profile"],
                "aperture_profile": row["aperture_profile"],
                "structural_signature": structural_signature,
                "release_year": None,
                "designer": None,
                "foundry": None,
                "license_type": "unknown",
                "license_url": None,
                "fallback_stack": None,
                "expert_enabled": False,
                "min_mode": "training",
                "qa_status": "review",
                "notes": notes,
            }
        )

    return records


def build_runtime_records(manifest: dict, project_root: Path) -> list[dict]:
    records: list[dict] = []
    generated_at = datetime.now(timezone.utc).isoformat()

    for font in manifest.get("fonts", []):
        slug = font["slug"]
        asset_status = font.get("assetStatus")
        runtime_path = font.get("runtimePath")

        if asset_status == "system_local":
            records.append(
                {
                    "typeface_slug": slug,
                    "file_role": "primary",
                    "font_format": "woff2",
                    "weight": 400,
                    "style": "normal",
                    "source_path": "system_local",
                    "runtime_path": None,
                    "file_size_bytes": None,
                    "sha256_hash": None,
                    "runtime_status": "system_local",
                    "asset_origin": "system_local",
                    "verified_at_utc": None,
                    "notes": [
                        "system font only",
                        "requires replacement or mirrored libre runtime asset before launch",
                    ],
                }
            )
            continue

        if not runtime_path:
            records.append(
                {
                    "typeface_slug": slug,
                    "file_role": "primary",
                    "font_format": "woff2",
                    "weight": 400,
                    "style": "normal",
                    "source_path": None,
                    "runtime_path": None,
                    "file_size_bytes": None,
                    "sha256_hash": None,
                    "runtime_status": "missing",
                    "asset_origin": asset_status or "unknown",
                    "verified_at_utc": None,
                    "notes": ["manifest has no runtimePath for this typeface"],
                }
            )
            continue

        runtime_file = project_root / "public" / runtime_path.lstrip("/")
        exists = runtime_file.exists()
        records.append(
            {
                "typeface_slug": slug,
                "file_role": "primary",
                "font_format": "woff2",
                "weight": 400,
                "style": "normal",
                "source_path": str(runtime_file.relative_to(project_root)) if exists else str(runtime_file),
                "runtime_path": runtime_path,
                "file_size_bytes": runtime_file.stat().st_size if exists else None,
                "sha256_hash": sha256_file(runtime_file) if exists else None,
                "runtime_status": "ready" if exists else "missing",
                "asset_origin": "google_fonts_mirror",
                "verified_at_utc": generated_at if exists else None,
                "notes": [] if exists else ["runtime file missing from public/fonts"],
            }
        )

    return records


def build_expert_answer_records(typeface_records: list[dict]) -> list[dict]:
    records: list[dict] = []
    for row in typeface_records:
        records.append(
            {
                "typeface_slug": row["typeface_slug"],
                "answer_text": row["display_name"],
                "answer_normalized": normalize_ascii(str(row["display_name"])),
                "is_canonical": True,
                "locale": "any",
                "qa_status": "draft",
                "notes": ["generated_from_display_name"],
            }
        )
    return records


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")


def main() -> None:
    args = parse_args()

    project_root = Path.cwd()
    excel_path = Path(args.excel).expanduser().resolve()
    manifest_path = (project_root / args.manifest).resolve() if not Path(args.manifest).is_absolute() else Path(args.manifest).resolve()
    output_dir = (project_root / args.output_dir).resolve() if not Path(args.output_dir).is_absolute() else Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_rows = load_excel_rows(excel_path, args.sheet)
    manifest = load_manifest(manifest_path)

    typeface_records = build_typeface_records(excel_rows)
    runtime_records = build_runtime_records(manifest, project_root)
    expert_records = build_expert_answer_records(typeface_records)

    generated_at = datetime.now(timezone.utc).isoformat()
    meta_payload = {
        "generated_at": generated_at,
        "status": "seed",
        "sources": {
            "excel": str(excel_path),
            "manifest": str(manifest_path),
            "sheet": args.sheet,
        },
        "counts": {
            "typefaces_total": len(typeface_records),
            "typefaces_active": sum(1 for item in typeface_records if item["activation_status"]),
            "runtime_assets_total": len(runtime_records),
            "runtime_assets_ready": sum(1 for item in runtime_records if item["runtime_status"] == "ready"),
            "expert_answer_keys_total": len(expert_records),
        },
        "notes": [
            "This is the canonical V1 seed format inside the repo.",
            "Editorial enrichment and review are expected after generation.",
        ],
    }

    write_json(output_dir / "catalog-meta.json", meta_payload)
    write_json(
        output_dir / "typefaces-core.seed.json",
        {"meta": {"generated_at": generated_at, "record_count": len(typeface_records)}, "records": typeface_records},
    )
    write_json(
        output_dir / "font-runtime-assets.seed.json",
        {"meta": {"generated_at": generated_at, "record_count": len(runtime_records)}, "records": runtime_records},
    )
    write_json(
        output_dir / "expert-answer-keys.seed.json",
        {"meta": {"generated_at": generated_at, "record_count": len(expert_records)}, "records": expert_records},
    )

    print(json.dumps(meta_payload, indent=2, ensure_ascii=True))


if __name__ == "__main__":
    main()
