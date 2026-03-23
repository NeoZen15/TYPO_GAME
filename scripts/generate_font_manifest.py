#!/usr/bin/env python3
"""Generate a typeface manifest from the validated Excel catalogue + WOFF2 folder.

Usage:
  python3 scripts/generate_font_manifest.py \
    --catalogue "/abs/path/JEUX_DE_TYPO_catalogue_v4__VALIDE__20260219_1903.xlsx" \
    --woff2-dir "/abs/path/01_woff2" \
    --output "content/typefaces/font-manifest-v4.json"
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SHEET = "typefaces"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a JSON typeface manifest from catalogue + WOFF2 assets."
    )
    parser.add_argument("--catalogue", required=True, help="Path to validated .xlsx catalogue")
    parser.add_argument("--woff2-dir", required=True, help="Directory containing .woff2 files")
    parser.add_argument(
        "--output",
        default="content/typefaces/font-manifest-v4.json",
        help="Output JSON path inside the project",
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Workbook sheet name")
    return parser.parse_args()


def parse_json_or_keep(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return None
    if not (stripped.startswith("{") or stripped.startswith("[")):
        return value
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        return value


def boolish(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "active"}
    return bool(value)


def main() -> None:
    args = parse_args()
    catalogue_path = Path(args.catalogue).expanduser().resolve()
    woff2_dir = Path(args.woff2_dir).expanduser().resolve()
    output_path = Path(args.output).resolve()

    if not catalogue_path.exists():
        raise SystemExit(f"Catalogue file not found: {catalogue_path}")
    if not woff2_dir.exists():
        raise SystemExit(f"WOFF2 directory not found: {woff2_dir}")

    workbook = load_workbook(catalogue_path, read_only=True, data_only=True)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(
            f"Sheet '{args.sheet}' not found. Available: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[args.sheet]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Sheet '{args.sheet}' is empty")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    required_headers = [
        "typeface_slug",
        "display_name",
        "primary_category",
        "sub_category",
        "difficulty_base",
        "rarity_tag",
        "visual_cluster_id",
        "dreyfus_tier",
        "activation_status",
        "font_source",
        "is_variable_font",
        "files_manifest",
        "year_tag",
        "weight_structure",
        "contrast_profile",
        "aperture_profile",
        "structural_signature",
    ]
    missing_headers = [name for name in required_headers if name not in header_index]
    if missing_headers:
        raise SystemExit(f"Missing required headers in catalogue: {', '.join(missing_headers)}")

    woff2_files = sorted([path.name for path in woff2_dir.glob("*.woff2")])
    by_slug: dict[str, list[str]] = {}
    for filename in woff2_files:
        # Naming convention observed: "{slug}__{hash}.woff2"
        slug = filename.split("__", 1)[0].strip()
        if not slug:
            continue
        by_slug.setdefault(slug, []).append(filename)

    manifest_fonts: list[dict[str, Any]] = []
    missing_asset_slugs: list[str] = []
    local_system_slugs: list[str] = []

    data_rows = [row for row in rows[1:] if any(cell is not None and str(cell).strip() for cell in row)]
    for row in data_rows:
        row_dict = {
            key: row[idx] if idx < len(row) else None for key, idx in header_index.items()
        }
        slug = str(row_dict["typeface_slug"]).strip()
        if not slug:
            continue

        files = by_slug.get(slug, [])
        font_source = str(row_dict["font_source"]).strip() if row_dict["font_source"] else None

        if not files:
            if font_source == "local":
                asset_status = "system_local"
                local_system_slugs.append(slug)
            else:
                asset_status = "missing_woff2"
                missing_asset_slugs.append(slug)
        else:
            asset_status = "mapped"

        manifest_fonts.append(
            {
                "slug": slug,
                "displayName": row_dict["display_name"],
                "primaryCategory": row_dict["primary_category"],
                "subCategory": row_dict["sub_category"],
                "difficultyBase": row_dict["difficulty_base"],
                "rarityTag": row_dict["rarity_tag"],
                "visualClusterId": row_dict["visual_cluster_id"],
                "dreyfusTier": row_dict["dreyfus_tier"],
                "activationStatus": boolish(row_dict["activation_status"]),
                "fontSource": font_source,
                "isVariableFont": boolish(row_dict["is_variable_font"]),
                "yearTag": row_dict["year_tag"],
                "weightStructure": row_dict["weight_structure"],
                "contrastProfile": row_dict["contrast_profile"],
                "apertureProfile": row_dict["aperture_profile"],
                "structuralSignature": parse_json_or_keep(row_dict["structural_signature"]),
                "filesManifestFromCatalogue": parse_json_or_keep(row_dict["files_manifest"]),
                "assetStatus": asset_status,
                "woff2Files": files,
            }
        )

    activation_counter = Counter(font["activationStatus"] for font in manifest_fonts)
    category_counter = Counter(font["primaryCategory"] for font in manifest_fonts)
    source_counter = Counter(font["fontSource"] for font in manifest_fonts)

    payload = {
        "meta": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceCatalogue": str(catalogue_path),
            "sourceWoff2Dir": str(woff2_dir),
            "sheet": args.sheet,
            "generator": "scripts/generate_font_manifest.py",
        },
        "summary": {
            "catalogueRows": len(manifest_fonts),
            "activeRows": activation_counter.get(True, 0),
            "inactiveRows": activation_counter.get(False, 0),
            "woff2FilesTotal": len(woff2_files),
            "mappedTypefaces": sum(1 for font in manifest_fonts if font["assetStatus"] == "mapped"),
            "systemLocalTypefaces": len(local_system_slugs),
            "missingWoff2Typefaces": len(missing_asset_slugs),
            "primaryCategoryBreakdown": dict(category_counter),
            "fontSourceBreakdown": dict(source_counter),
        },
        "missingWoff2Slugs": sorted(missing_asset_slugs),
        "systemLocalSlugs": sorted(local_system_slugs),
        "fonts": manifest_fonts,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Manifest generated: {output_path}")
    print(f"Typefaces: {payload['summary']['catalogueRows']}")
    print(f"Mapped with WOFF2: {payload['summary']['mappedTypefaces']}")
    print(f"System local only: {payload['summary']['systemLocalTypefaces']}")
    print(f"Missing WOFF2 (unexpected): {payload['summary']['missingWoff2Typefaces']}")


if __name__ == "__main__":
    main()
